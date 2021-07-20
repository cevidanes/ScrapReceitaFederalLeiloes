import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import os


#quando não quiser validar editar existentes trocar para N. Se Y ele pergunta a cada edital que existir se deseja sobescrever.
questionar_existentes="Y"

# PrimeiraPagina:
# http://www25.receita.fazenda.gov.br/sle-sociedade/api/portal
# Load in the workbook

def cria_planilha_edital(cod_edital, data_fim_proposta, edital):
    nome_planilha = cod_edital.replace('/', '_') + "_" + data_fim_proposta.replace(' ', '_').replace(':', '_') + '.xlsx'
    print(edital)

    if questionar_existentes=="Y":
        #Verificar se a planilha existe
        if os.path.isfile(nome_planilha):
            txt = input(
                f"Encontrei um arquivo gerado para o lote com nome {nome_planilha} se deseja importar mesmo assim digite ENTER, caso contrário digite N para prosseguir:")

            if txt.upper() == "N":
                return "N"

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalhes Edital"

    ws["A1"] = "Nome Edital"
    ws["A2"] = edital['edital'] + ' - ' + edital['cidade']
    ws["B1"] = "Data Inicio Propostas"
    ws["B2"] = edital['dataInicioPropostas']
    ws["C1"] = "Data Fim Propostas"
    ws["C2"] = edital['dataFimPropostas']
    ws["D1"] = "Data Abertura de Lances"
    ws["D2"] = edital['dataAberturaLances']
    ws["E1"] = "Número de Lotes"
    ws["E2"] = edital['lotes']
    ws["F1"] = "Link Edital"
    ws["F2"] = "http://www25.receita.fazenda.gov.br/sle-sociedade/portal/edital/" + edital['edle']
    wb.save(nome_planilha)

    wb.close()
    return nome_planilha


def cria_planilha_lote(edital_database, sValorMinimo, sErratas, lista_items, i):
    wb = load_workbook(filename=edital_database)
    sheet_ranges = wb['Detalhes Edital']
    # print(sheet_ranges['A2'].value)
    # print(str(i))
    ws2 = wb.create_sheet(title=str(i))

    ws2["A1"] = "Local Armazenamento"
    ws2["B1"] = "Quantidade"
    ws2["C1"] = "Unidade"
    ws2["D1"] = "Descricao"
    ws2["E1"] = "Valor Unitário"
    ws2["F1"] = "Valor Total"
    ws2["G1"] = "Valor Minimo Lote"

    ws2["K3"] = "Valor Total do Lote em Vendas"  # =SOMA(F3:F5)
    ws2["K4"] = "Lance maximo"
    ws2["K5"] = "ICMS 25% do Lance"  # =SOMA(B7)*25%
    ws2["K6"] = "Frete"
    ws2["K7"] = "Custo Total do Lote"  # =SOMA(B7+B8+B9)
    ws2["K8"] = "Lucro"  # =SOMA(B6-B10)
    ws2["K9"] = "Lucro ideal seria 60% do custo total"  # =SOMA(B10)*60%

    index = 1
    for item in lista_items:
        index = index + 1
        ws2["A" + str(index)] = str(item[0])
        ws2["B" + str(index)] = str(item[1])
        ws2["C" + str(index)] = str(item[2])
        ws2["D" + str(index)] = str(item[3])
        ws2["G" + str(index)] = sValorMinimo
        ws2["F" + str(index)] = "=B" + str(index) + "*" + "E" + str(index)

    sFormulaValortotalVendas = "=SOMA(F2:F" + str(index + 1) + ")"
    ws2["L3"] = sFormulaValortotalVendas
    sICSMS = "=SOMA(L4)*25%"
    ws2["L5"] = sICSMS
    sCustoTotal = "=SOMA(L6+L5+L4)"
    ws2["L7"] = sCustoTotal
    sLucro = "=SOMA(L3-L7)"
    ws2["L8"] = sLucro
    sLucroIDeal = "=SOMA(L7)*60%"
    ws2["L9"] = sLucroIDeal

    wb.save(edital_database)
    wb.close()

    pass


def consulta_lotes_e_gera_planilha():
    # print('Chama essa função!')
    URL = 'http://www25.receita.fazenda.gov.br/sle-sociedade/api/portal'
    data = requests.get(URL)
    retorno = data.json()

    for item in retorno['situacoes']:
        print(item['label'])
        # EDITAL
        for edital in item['lista']:

            edital_database = cria_planilha_edital(edital['edital'] + ' - ' + edital['cidade'],
                                                   edital['dataFimPropostas'], edital)

            if edital_database != "N":
                x = range(int(edital['lotes']))
                # LOTE
                for i in x:
                    i = i + 1
                    URL = 'http://www25.receita.fazenda.gov.br/sle-sociedade/api/lote/' + edital['edle'] + '/' + str(i)
                    data = requests.get(URL)
                    retorno = data.json()
                    sValorMinimo = retorno['valorMinimo']
                    # print(sValorMinimo)
                    sErratas = ""
                    if 'avisosErratas' in retorno.keys():
                        sErratas = retorno['avisosErratas']

                    lista_items = []
                    # ITEMS DO LOTE
                    for item in retorno['itensDetalhesLote']:
                        # ITEM DO LOTE
                        item = (item['recintoArmazenador'], item['quantidade'], item['unMedida'], item['descricao'])
                        lista_items.append(item)

                    cria_planilha_lote(edital_database, sValorMinimo, sErratas, lista_items, i)


# Começa aqui
print('começa aqui!')
consulta_lotes_e_gera_planilha()
